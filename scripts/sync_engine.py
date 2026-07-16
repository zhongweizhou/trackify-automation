"""Incrementally synchronize Excel-managed scenarios into Gherkin features."""

from __future__ import annotations

import argparse
import hashlib
import os
import re
import shutil
import subprocess
import sys
import tempfile
import threading
import time
from collections.abc import Callable, Sequence
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from watchdog.events import FileSystemEvent, FileSystemEventHandler
from watchdog.observers import Observer

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_INPUT = ROOT / "data" / "test_cases.xlsx"
BACKUP_DIR_NAME = ".backup"
LOCK_FILE_NAME = "sync.lock"
WATCH_DEBOUNCE_SECONDS = 5.0

MODULE_PATHS = {
    "add_transaction": Path("tests/features/add_transaction.feature"),
    "transactions": Path("tests/features/transactions.feature"),
}
STEP_MODULES = {
    "add_transaction": Path("tests/step_defs/add_transaction_steps.py"),
    "transactions": Path("tests/step_defs/transactions_steps.py"),
}
REQUIRED_HEADERS = (
    "Test Case ID",
    "Module",
    "Scenario Title",
    "Priority",
    "App Version Introduced",
    "App Version Deprecated",
    "Platform",
    "Automation Status",
    "Author",
    "Last Reviewed Date",
    "Last Run Result",
    "Tags",
    "Pre-conditions",
    "Test Steps",
    "Expected Result",
    "Estimated Runtime (s)",
)
ALLOWED_PRIORITIES = {"P0", "P1", "P2"}
ALLOWED_PLATFORMS = {"android", "ios", "both"}
ALLOWED_STATUSES = {"manual", "automated", "candidate", "deprecated"}
ALLOWED_RUN_RESULTS = {"pass", "fail", "skip", "not_run"}
ID_PATTERN = re.compile(r"^TC_[A-Z][A-Z0-9_]*_[0-9]{3}$")
SEMVER_PATTERN = re.compile(r"^[0-9]+\.[0-9]+\.[0-9]+(?:[-+][0-9A-Za-z.-]+)?$")
TAG_PATTERN = re.compile(r"^[a-z][a-z0-9_:-]*$")
STEP_KEYWORD_PATTERN = re.compile(r"^(?:Given|When|Then|And|But)\b", re.IGNORECASE)
BLOCK_START_PATTERN = re.compile(r"(?m)^# scenario_id:\s*(\S+)\s*$")
ACTIVE_SCENARIO_PATTERN = re.compile(r"(?m)^\s*Scenario:\s*(.+?)\s*$")
OUTLINE_PATTERN = re.compile(r"(?m)^\s*Scenario Outline:\s*")


class SyncError(Exception):
    """Base exception for a safe, user-actionable sync failure."""


class ValidationError(SyncError):
    """One or more workbook or feature validation errors."""

    def __init__(self, errors: Sequence[str]) -> None:
        self.errors = tuple(errors)
        super().__init__("\n".join(self.errors))


@dataclass(frozen=True)
class RegistryRow:
    """One validated Excel registry row."""

    row_number: int
    scenario_id: str
    module: str
    title: str
    priority: str
    introduced_in: str
    deprecated_in: str | None
    platform: str
    status: str
    author: str
    reviewed_date: str
    last_run_result: str
    tags: tuple[str, ...]
    preconditions: str
    steps: tuple[str, ...]
    expected: tuple[str, ...]
    estimated_runtime_seconds: int | None


@dataclass(frozen=True)
class ManagedBlock:
    """An exact managed byte slice within one feature document."""

    scenario_id: str
    start: int
    end: int
    text: str
    core: str
    trailing: str
    deprecated: bool


@dataclass(frozen=True)
class FeatureDocument:
    """A validated feature document with code-owned prefix and managed blocks."""

    module: str
    path: Path
    text: str
    newline: str
    blocks: tuple[ManagedBlock, ...]


@dataclass(frozen=True)
class Change:
    """One scenario-level change planned from the registry."""

    kind: str
    row: RegistryRow
    path: Path
    existing: ManagedBlock | None
    rendered_core: str


@dataclass(frozen=True)
class SyncPlan:
    """A fully validated, in-memory synchronization plan."""

    rows: tuple[RegistryRow, ...]
    documents: tuple[FeatureDocument, ...]
    changes: tuple[Change, ...]
    unchanged_ids: tuple[str, ...]
    rendered_documents: dict[Path, str]

    @property
    def has_drift(self) -> bool:
        return bool(self.changes)

    @property
    def changed_active_rows(self) -> tuple[RegistryRow, ...]:
        return tuple(
            change.row
            for change in self.changes
            if change.row.status == "automated"
        )


def file_sha256(path: Path) -> str:
    """Return a file SHA-256 without modifying it."""
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _iso_date(value: Any, row_number: int) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    raw = _text(value)
    try:
        return date.fromisoformat(raw).isoformat()
    except ValueError as exc:
        raise ValueError(
            f"row {row_number}: Last Reviewed Date must be ISO YYYY-MM-DD"
        ) from exc


def _phrases(value: Any, field: str, row_number: int) -> tuple[str, ...]:
    raw = "" if value is None else str(value).replace("\r\n", "\n").replace("\r", "\n")
    if not raw.strip():
        return ()
    lines = raw.split("\n")
    if any(not line.strip() for line in lines):
        raise ValueError(f"row {row_number}: {field} contains a blank phrase")
    phrases = tuple(line.strip() for line in lines)
    for phrase in phrases:
        if STEP_KEYWORD_PATTERN.match(phrase):
            raise ValueError(
                f"row {row_number}: {field} phrases must omit Gherkin keywords: "
                f"{phrase!r}"
            )
    return phrases


def _normalized_tags(value: Any, priority: str, row_number: int) -> tuple[str, ...]:
    raw_tags = [tag.strip().lower() for tag in _text(value).split(",") if tag.strip()]
    if any(tag.startswith("@") or not TAG_PATTERN.fullmatch(tag) for tag in raw_tags):
        raise ValueError(f"row {row_number}: Tags contain an invalid marker")

    canonical_priority = priority.lower()
    seen: set[str] = set()
    deduplicated: list[str] = []
    for tag in raw_tags:
        if tag not in seen:
            seen.add(tag)
            deduplicated.append(tag)
    priority_positions = [
        index for index, tag in enumerate(deduplicated) if tag in {"p0", "p1", "p2"}
    ]
    insert_at = priority_positions[0] if priority_positions else 0
    non_priorities = [tag for tag in deduplicated if tag not in {"p0", "p1", "p2"}]
    result = list(non_priorities)
    result.insert(min(insert_at, len(result)), canonical_priority)
    return tuple(result)


def _parse_registry_row(values: dict[str, Any], row_number: int) -> RegistryRow:
    scenario_id = _text(values["Test Case ID"])
    module = _text(values["Module"])
    title = _text(values["Scenario Title"])
    priority = _text(values["Priority"]).upper()
    introduced = _text(values["App Version Introduced"])
    deprecated = _text(values["App Version Deprecated"]) or None
    platform = _text(values["Platform"]).lower()
    status = _text(values["Automation Status"]).lower()
    author = _text(values["Author"])
    last_run_result = _text(values["Last Run Result"]).lower()

    required = {
        "Test Case ID": scenario_id,
        "Module": module,
        "Scenario Title": title,
        "Priority": priority,
        "App Version Introduced": introduced,
        "Platform": platform,
        "Automation Status": status,
        "Author": author,
        "Last Run Result": last_run_result,
    }
    missing = [name for name, value in required.items() if not value]
    if missing:
        raise ValueError(f"row {row_number}: missing required values: {', '.join(missing)}")
    if not ID_PATTERN.fullmatch(scenario_id):
        raise ValueError(f"row {row_number}: invalid Test Case ID {scenario_id!r}")
    if module not in MODULE_PATHS:
        raise ValueError(f"row {row_number}: unknown Module {module!r}")
    if priority not in ALLOWED_PRIORITIES:
        raise ValueError(f"row {row_number}: invalid Priority {priority!r}")
    if not SEMVER_PATTERN.fullmatch(introduced):
        raise ValueError(f"row {row_number}: invalid App Version Introduced")
    if deprecated and not SEMVER_PATTERN.fullmatch(deprecated):
        raise ValueError(f"row {row_number}: invalid App Version Deprecated")
    if platform not in ALLOWED_PLATFORMS:
        raise ValueError(f"row {row_number}: invalid Platform {platform!r}")
    if status not in ALLOWED_STATUSES:
        raise ValueError(f"row {row_number}: invalid Automation Status {status!r}")
    if last_run_result not in ALLOWED_RUN_RESULTS:
        raise ValueError(f"row {row_number}: invalid Last Run Result")
    if status == "deprecated" and deprecated is None:
        raise ValueError(
            f"row {row_number}: deprecated status requires App Version Deprecated"
        )

    steps = _phrases(values["Test Steps"], "Test Steps", row_number)
    expected = _phrases(values["Expected Result"], "Expected Result", row_number)
    if status == "automated" and (not steps or not expected):
        raise ValueError(
            f"row {row_number}: automated cases require Test Steps and Expected Result"
        )

    runtime_value = values["Estimated Runtime (s)"]
    runtime: int | None = None
    if runtime_value not in (None, ""):
        if isinstance(runtime_value, bool):
            raise ValueError(f"row {row_number}: Estimated Runtime must be positive")
        try:
            runtime_number = float(runtime_value)
        except (TypeError, ValueError) as exc:
            raise ValueError(
                f"row {row_number}: Estimated Runtime must be numeric"
            ) from exc
        if runtime_number <= 0 or not runtime_number.is_integer():
            raise ValueError(
                f"row {row_number}: Estimated Runtime must be a positive integer"
            )
        runtime = int(runtime_number)

    return RegistryRow(
        row_number=row_number,
        scenario_id=scenario_id,
        module=module,
        title=title,
        priority=priority,
        introduced_in=introduced,
        deprecated_in=deprecated,
        platform=platform,
        status=status,
        author=author,
        reviewed_date=_iso_date(values["Last Reviewed Date"], row_number),
        last_run_result=last_run_result,
        tags=_normalized_tags(values["Tags"], priority, row_number),
        preconditions=_text(values["Pre-conditions"]),
        steps=steps,
        expected=expected,
        estimated_runtime_seconds=runtime,
    )


def parse_workbook(path: Path) -> tuple[RegistryRow, ...]:
    """Parse and validate the registry without writing to it."""
    if not path.is_file():
        raise ValidationError([f"workbook not found: {path}"])
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValidationError([f"cannot open workbook {path}: {exc}"]) from exc
    try:
        sheet = workbook["Test Cases"] if "Test Cases" in workbook.sheetnames else workbook.active
        header_values = [_text(cell.value) for cell in sheet[1]]
        if len(header_values) != len(set(header_values)):
            raise ValidationError(["workbook contains duplicate headers"])
        missing_headers = [name for name in REQUIRED_HEADERS if name not in header_values]
        if missing_headers:
            raise ValidationError(
                [f"workbook missing headers: {', '.join(missing_headers)}"]
            )
        errors: list[str] = []
        rows: list[RegistryRow] = []
        for row_number, cells in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            values = dict(zip(header_values, cells))
            if not any(value not in (None, "") for value in cells):
                continue
            try:
                rows.append(_parse_registry_row(values, row_number))
            except ValueError as exc:
                errors.append(str(exc))

        ids: dict[str, int] = {}
        titles: dict[tuple[str, str], int] = {}
        for row in rows:
            if row.scenario_id in ids:
                errors.append(
                    f"row {row.row_number}: duplicate Test Case ID {row.scenario_id}; "
                    f"first seen on row {ids[row.scenario_id]}"
                )
            ids[row.scenario_id] = row.row_number
            title_key = (row.module, row.title.casefold())
            if title_key in titles:
                errors.append(
                    f"row {row.row_number}: duplicate Scenario Title in module "
                    f"{row.module!r}; first seen on row {titles[title_key]}"
                )
            titles[title_key] = row.row_number
        if errors:
            raise ValidationError(errors)
        return tuple(rows)
    finally:
        workbook.close()


def validate_registry_rows(rows: Sequence[RegistryRow]) -> None:
    """Validate row objects before routing or rendering feature paths."""
    errors: list[str] = []
    ids: dict[str, int] = {}
    titles: dict[tuple[str, str], int] = {}
    for row in rows:
        prefix = f"row {row.row_number}"
        if not ID_PATTERN.fullmatch(row.scenario_id):
            errors.append(f"{prefix}: invalid Test Case ID {row.scenario_id!r}")
        if row.module not in MODULE_PATHS:
            errors.append(f"{prefix}: unknown Module {row.module!r}")
        if not row.title:
            errors.append(f"{prefix}: missing Scenario Title")
        if row.priority not in ALLOWED_PRIORITIES:
            errors.append(f"{prefix}: invalid Priority {row.priority!r}")
        if row.platform not in ALLOWED_PLATFORMS:
            errors.append(f"{prefix}: invalid Platform {row.platform!r}")
        if row.status not in ALLOWED_STATUSES:
            errors.append(f"{prefix}: invalid Automation Status {row.status!r}")
        if row.last_run_result not in ALLOWED_RUN_RESULTS:
            errors.append(f"{prefix}: invalid Last Run Result")
        if row.status == "automated" and (not row.steps or not row.expected):
            errors.append(f"{prefix}: automated case lacks steps or expected results")
        if row.status == "deprecated" and not row.deprecated_in:
            errors.append(f"{prefix}: deprecated case lacks deprecated version")
        if row.scenario_id in ids:
            errors.append(
                f"{prefix}: duplicate Test Case ID {row.scenario_id}; "
                f"first seen on row {ids[row.scenario_id]}"
            )
        ids[row.scenario_id] = row.row_number
        title_key = (row.module, row.title.casefold())
        if title_key in titles:
            errors.append(
                f"{prefix}: duplicate Scenario Title in module {row.module!r}; "
                f"first seen on row {titles[title_key]}"
            )
        titles[title_key] = row.row_number
    if errors:
        raise ValidationError(errors)


def _detect_newline(text: str, path: Path) -> str:
    has_crlf = "\r\n" in text
    has_lf = "\n" in text.replace("\r\n", "")
    if has_crlf and has_lf:
        raise ValidationError([f"mixed newline styles in {path}"])
    return "\r\n" if has_crlf else "\n"


def parse_feature(module: str, path: Path) -> FeatureDocument:
    """Parse managed blocks while preserving their exact source slices."""
    try:
        text = path.read_bytes().decode("utf-8")
    except (OSError, UnicodeDecodeError) as exc:
        raise ValidationError([f"cannot read UTF-8 feature {path}: {exc}"]) from exc
    newline = _detect_newline(text, path)
    if OUTLINE_PATTERN.search(text):
        raise ValidationError([f"Scenario Outline is unsupported in {path}"])

    starts = list(BLOCK_START_PATTERN.finditer(text))
    blocks: list[ManagedBlock] = []
    errors: list[str] = []
    seen: set[str] = set()
    for index, match in enumerate(starts):
        start = match.start()
        end = starts[index + 1].start() if index + 1 < len(starts) else len(text)
        raw = text[start:end]
        core = raw.rstrip("\r\n")
        trailing = raw[len(core) :]
        scenario_id = match.group(1)
        if scenario_id in seen:
            errors.append(f"duplicate feature ID {scenario_id} in {path}")
        seen.add(scenario_id)
        deprecated = f"# DEPRECATED BEGIN {scenario_id}" in core
        scenario_matches = list(ACTIVE_SCENARIO_PATTERN.finditer(core))
        expected_scenarios = 0 if deprecated else 1
        if len(scenario_matches) != expected_scenarios:
            errors.append(
                f"managed block {scenario_id} in {path} contains "
                f"{len(scenario_matches)} active Scenario lines; expected "
                f"{expected_scenarios} (possible unmanaged Scenario)"
            )
        blocks.append(
            ManagedBlock(
                scenario_id=scenario_id,
                start=start,
                end=end,
                text=raw,
                core=core,
                trailing=trailing,
                deprecated=deprecated,
            )
        )

    for scenario in ACTIVE_SCENARIO_PATTERN.finditer(text):
        if not any(block.start <= scenario.start() < block.end for block in blocks):
            errors.append(
                f"unmanaged Scenario {scenario.group(1)!r} in {path}; add scenario_id metadata"
            )
    if errors:
        raise ValidationError(errors)
    return FeatureDocument(
        module=module,
        path=path,
        text=text,
        newline=newline,
        blocks=tuple(blocks),
    )


def _platforms(platform: str) -> str:
    return "android, ios" if platform == "both" else platform


def _render_active(row: RegistryRow, newline: str) -> str:
    lines = [
        f"# scenario_id: {row.scenario_id}",
        f"# introduced_in: {row.introduced_in}",
        f"# platforms: {_platforms(row.platform)}",
        "  " + " ".join(f"@{tag}" for tag in row.tags),
        f"  Scenario: {row.title}",
    ]
    for index, phrase in enumerate(row.steps):
        lines.append(f"    {'When' if index == 0 else 'And'} {phrase}")
    for index, phrase in enumerate(row.expected):
        lines.append(f"    {'Then' if index == 0 else 'And'} {phrase}")
    return newline.join(lines)


def _existing_active_body(block: ManagedBlock) -> list[str]:
    lines = block.core.splitlines()
    if block.deprecated:
        begin = next(
            index
            for index, line in enumerate(lines)
            if line == f"# DEPRECATED BEGIN {block.scenario_id}"
        )
        end = next(
            index
            for index, line in enumerate(lines)
            if line == f"# DEPRECATED END {block.scenario_id}"
        )
        return [
            line[2:] if line.startswith("# ") else line[1:]
            for line in lines[begin + 1 : end]
        ]
    tag_index = next(
        (index for index, line in enumerate(lines) if line.lstrip().startswith("@")),
        None,
    )
    if tag_index is None:
        raise ValidationError(
            [f"managed block {block.scenario_id} has no executable body"]
        )
    return lines[tag_index:]


def _render_deprecated(
    row: RegistryRow,
    newline: str,
    existing: ManagedBlock,
) -> str:
    if row.steps and row.expected:
        active_lines = _render_active(row, newline).splitlines()[3:]
    else:
        active_lines = _existing_active_body(existing)
    lines = [
        f"# scenario_id: {row.scenario_id}",
        f"# introduced_in: {row.introduced_in}",
        f"# deprecated_in: {row.deprecated_in}",
        f"# platforms: {_platforms(row.platform)}",
        f"# DEPRECATED BEGIN {row.scenario_id}",
        *(f"# {line}" for line in active_lines),
        f"# DEPRECATED END {row.scenario_id}",
    ]
    return newline.join(lines)


def _render_row(
    row: RegistryRow,
    document: FeatureDocument,
    existing: ManagedBlock | None,
) -> str:
    if row.status == "deprecated":
        if existing is None:
            raise ValidationError(
                [f"cannot deprecate unknown scenario {row.scenario_id}"]
            )
        return _render_deprecated(row, document.newline, existing)
    return _render_active(row, document.newline)


def _render_document(
    document: FeatureDocument,
    changes: Sequence[Change],
) -> str:
    existing_replacements = {
        change.row.scenario_id: change.rendered_core
        for change in changes
        if change.existing is not None
    }
    rendered: list[str] = []
    cursor = 0
    for block in document.blocks:
        rendered.append(document.text[cursor : block.start])
        replacement = existing_replacements.get(block.scenario_id)
        rendered.append(
            block.text if replacement is None else replacement + block.trailing
        )
        cursor = block.end
    rendered.append(document.text[cursor:])
    text = "".join(rendered)

    additions = [change.rendered_core for change in changes if change.existing is None]
    if additions:
        if text and not text.endswith(document.newline):
            text += document.newline
        text += document.newline.join(additions) + document.newline
    return text


def build_plan(input_path: Path, root: Path = ROOT) -> SyncPlan:
    """Validate all inputs and build a no-write synchronization plan."""
    rows = parse_workbook(input_path)
    return build_plan_from_rows(rows, root)


def build_plan_from_rows(
    rows: Sequence[RegistryRow],
    root: Path = ROOT,
) -> SyncPlan:
    """Build a synchronization plan from already validated registry rows."""
    validate_registry_rows(rows)
    documents = tuple(
        parse_feature(module, root / relative_path)
        for module, relative_path in MODULE_PATHS.items()
    )
    document_by_module = {document.module: document for document in documents}
    block_locations: dict[str, tuple[FeatureDocument, ManagedBlock]] = {}
    errors: list[str] = []
    for document in documents:
        for block in document.blocks:
            if block.scenario_id in block_locations:
                first = block_locations[block.scenario_id][0].path
                errors.append(
                    f"duplicate feature ID {block.scenario_id} across {first} and {document.path}"
                )
            block_locations[block.scenario_id] = (document, block)

    row_by_id = {row.scenario_id: row for row in rows}
    for scenario_id, (document, _) in block_locations.items():
        row = row_by_id.get(scenario_id)
        if row is None:
            errors.append(
                f"managed feature ID {scenario_id} in {document.path} is missing from workbook"
            )
        elif row.status in {"manual", "candidate"}:
            errors.append(
                f"{scenario_id} exists in a feature but workbook status is {row.status}; "
                "transition it to deprecated explicitly"
            )

    changes: list[Change] = []
    unchanged: list[str] = []
    for row in rows:
        location = block_locations.get(row.scenario_id)
        if row.status in {"manual", "candidate"}:
            continue
        document = document_by_module[row.module]
        existing = location[1] if location else None
        if location and location[0].module != row.module:
            errors.append(
                f"{row.scenario_id} is routed to {row.module} but exists in "
                f"{location[0].module}"
            )
            continue
        if row.status == "deprecated" and existing is None:
            errors.append(f"cannot deprecate unknown scenario {row.scenario_id}")
            continue
        try:
            rendered_core = _render_row(row, document, existing)
        except ValidationError as exc:
            errors.extend(exc.errors)
            continue
        if existing is None:
            kind = "added"
        elif existing.core == rendered_core:
            unchanged.append(row.scenario_id)
            continue
        elif row.status == "deprecated" and not existing.deprecated:
            kind = "deprecated"
        else:
            kind = "modified"
        changes.append(
            Change(
                kind=kind,
                row=row,
                path=document.path,
                existing=existing,
                rendered_core=rendered_core,
            )
        )
    if errors:
        raise ValidationError(errors)

    rendered_documents: dict[Path, str] = {}
    for document in documents:
        document_changes = [change for change in changes if change.path == document.path]
        if document_changes:
            rendered_documents[document.path] = _render_document(
                document,
                document_changes,
            )
    return SyncPlan(
        rows=tuple(rows),
        documents=documents,
        changes=tuple(changes),
        unchanged_ids=tuple(sorted(unchanged)),
        rendered_documents=rendered_documents,
    )


def print_plan(plan: SyncPlan, root: Path = ROOT) -> None:
    """Print one deterministic, scenario-attributable plan summary."""
    for module, relative_path in MODULE_PATHS.items():
        module_changes = [change for change in plan.changes if change.row.module == module]
        counts = {
            kind: sum(change.kind == kind for change in module_changes)
            for kind in ("added", "modified", "deprecated")
        }
        unchanged = sum(
            row.module == module and row.scenario_id in plan.unchanged_ids
            for row in plan.rows
        )
        print(
            f"[sync] {relative_path}: added={counts['added']} "
            f"modified={counts['modified']} deprecated={counts['deprecated']} "
            f"unchanged={unchanged}"
        )
        for change in module_changes:
            print(
                f"  - {change.kind}: {change.row.scenario_id} "
                f"({change.row.title})"
            )
    if plan.has_drift:
        print("[sync] Drift detected. Run with --apply to update managed blocks.")
    else:
        print("[sync] Registry and managed feature blocks are in sync.")


def _atomic_write(path: Path, content: bytes) -> None:
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.",
        suffix=".tmp",
        dir=path.parent,
    )
    temporary_path = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(content)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary_path, path)
    finally:
        temporary_path.unlink(missing_ok=True)


def _default_collection(root: Path) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        [sys.executable, "-m", "pytest", "--collect-only", "-q"],
        cwd=root,
        check=False,
        capture_output=True,
        text=True,
    )


def apply_plan(
    plan: SyncPlan,
    workbook_path: Path,
    root: Path = ROOT,
    collection_runner: Callable[[Path], subprocess.CompletedProcess[str]] | None = None,
) -> None:
    """Atomically apply a validated plan and roll back collection failures."""
    if not plan.has_drift:
        return
    backup_dir = root / "data" / BACKUP_DIR_NAME
    backup_dir.mkdir(parents=True, exist_ok=True)
    lock_path = backup_dir / LOCK_FILE_NAME
    try:
        lock_descriptor = os.open(lock_path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
    except FileExistsError as exc:
        raise SyncError(
            f"sync lock exists: {lock_path}; confirm its PID is inactive before removal"
        ) from exc
    backups: dict[Path, Path] = {}
    try:
        with os.fdopen(lock_descriptor, "w", encoding="utf-8") as lock_handle:
            lock_handle.write(
                f"pid={os.getpid()}\n"
                f"started={datetime.now().astimezone().isoformat()}\n"
            )
            lock_handle.flush()
            os.fsync(lock_handle.fileno())

        workbook_hash = file_sha256(workbook_path)
        timestamp = datetime.now().strftime("%Y%m%dT%H%M%S%f")
        try:
            for path in plan.rendered_documents:
                backup = backup_dir / f"{path.name}.{timestamp}.bak"
                shutil.copy2(path, backup)
                backups[path] = backup
            for path, rendered in plan.rendered_documents.items():
                _atomic_write(path, rendered.encode("utf-8"))

            runner = collection_runner or _default_collection
            completed = runner(root)
            if completed.returncode != 0:
                detail = (completed.stdout + "\n" + completed.stderr).strip()
                raise SyncError(f"post-write pytest collection failed:\n{detail}")
            if file_sha256(workbook_path) != workbook_hash:
                raise SyncError("workbook changed during apply")
        except Exception as original_error:
            restore_errors: list[str] = []
            for path, backup in backups.items():
                try:
                    _atomic_write(path, backup.read_bytes())
                except Exception as restore_error:
                    restore_errors.append(f"{path}: {restore_error}")
            suffix = (
                f"; rollback errors: {'; '.join(restore_errors)}"
                if restore_errors
                else "; all changed features restored"
            )
            raise SyncError(f"{original_error}{suffix}") from original_error
    finally:
        lock_path.unlink(missing_ok=True)


def _python_name(title: str) -> str:
    value = re.sub(r"\W", "", title.replace(" ", "_"))
    return re.sub(r"^\d+_*", "", value).lower()


def changed_nodeids(rows: Sequence[RegistryRow]) -> tuple[str, ...]:
    """Return exact pytest node IDs for active changed scenarios."""
    return tuple(
        f"{STEP_MODULES[row.module]}::test_{_python_name(row.title)}"
        for row in rows
    )


def run_changed_cases(rows: Sequence[RegistryRow], root: Path = ROOT) -> int:
    """Run only active changed scenarios and print actionable feedback."""
    nodeids = changed_nodeids(rows)
    if not nodeids:
        print("[sync] No added or modified active scenarios require execution.")
        return 0
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    allure_dir = root / "report" / "sync" / timestamp / "allure-results"
    allure_dir.mkdir(parents=True, exist_ok=True)
    command = [
        sys.executable,
        "-m",
        "pytest",
        "-q",
        *nodeids,
        f"--alluredir={allure_dir}",
    ]
    print("[sync] Running changed scenarios only:")
    for row, nodeid in zip(rows, nodeids, strict=True):
        print(f"  - {row.scenario_id}: {nodeid}")
    completed = subprocess.run(command, cwd=root, check=False)
    if completed.returncode == 0:
        print(
            f"[sync] Changed-case execution PASSED: {len(nodeids)} scenario(s). "
            f"Allure results: {allure_dir}"
        )
        return 0
    print(
        f"[sync] Changed-case execution FAILED with exit {completed.returncode}.\n"
        "[sync] Feature changes were kept for debugging. Review Task 13 triage, "
        "failure screenshots, and the Allure results; add or update step definitions, "
        "Page/Flow behavior, and YAML locators using Appium evidence, then rerun:\n"
        f"[sync] {' '.join(command)}\n"
        f"[sync] Allure results: {allure_dir}"
    )
    return 1


def execute_once(
    input_path: Path,
    apply: bool,
    run_changed: bool,
    root: Path = ROOT,
) -> int:
    """Execute one check/apply cycle while proving the workbook is read-only."""
    before_hash = file_sha256(input_path)
    plan = build_plan(input_path, root)
    print_plan(plan, root)
    if not apply:
        if file_sha256(input_path) != before_hash:
            raise SyncError("workbook changed during check")
        return 1 if plan.has_drift else 0

    apply_plan(plan, input_path, root)
    if file_sha256(input_path) != before_hash:
        raise SyncError("workbook changed during apply")
    if plan.has_drift:
        print(f"[sync] Applied {len(plan.changes)} scenario change(s) successfully.")
    if run_changed:
        return run_changed_cases(plan.changed_active_rows, root)
    return 0


class _RegistryEventHandler(FileSystemEventHandler):
    def __init__(self, target: Path, pending: threading.Event) -> None:
        self.target = target.resolve()
        self.pending = pending

    def on_any_event(self, event: FileSystemEvent) -> None:
        if event.is_directory:
            return
        candidates = [Path(event.src_path).resolve()]
        destination = getattr(event, "dest_path", None)
        if destination:
            candidates.append(Path(destination).resolve())
        if self.target in candidates:
            self.pending.set()


def watch_registry(
    input_path: Path,
    apply: bool,
    run_changed: bool,
    root: Path = ROOT,
) -> int:
    """Watch the registry and invoke the same debounced one-shot path."""
    pending = threading.Event()
    observer = Observer()
    observer.schedule(
        _RegistryEventHandler(input_path, pending),
        str(input_path.parent),
        recursive=False,
    )
    observer.start()
    print(
        f"[sync] Watching {input_path} with {WATCH_DEBOUNCE_SECONDS:.0f}s debounce "
        f"({'apply' if apply else 'check'} mode). Press Ctrl-C to stop."
    )
    try:
        while True:
            pending.wait()
            pending.clear()
            while pending.wait(WATCH_DEBOUNCE_SECONDS):
                pending.clear()
            try:
                execute_once(input_path, apply, run_changed, root)
            except SyncError as exc:
                print(f"[sync] ERROR: {exc}", file=sys.stderr)
    except KeyboardInterrupt:
        print("\n[sync] Watch stopped.")
        return 0
    finally:
        observer.stop()
        observer.join()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Validate and incrementally sync Excel-managed BDD scenarios."
    )
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--check", action="store_true", help="Validate and report drift")
    mode.add_argument("--apply", action="store_true", help="Apply validated drift")
    parser.add_argument("--watch", action="store_true", help="Watch the input workbook")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument(
        "--run-changed",
        action="store_true",
        help="After apply, execute only added or modified active scenarios",
    )
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    if args.run_changed and not args.apply:
        print("[sync] ERROR: --run-changed requires --apply", file=sys.stderr)
        return 2
    input_path = args.input.expanduser().resolve()
    try:
        if args.watch:
            return watch_registry(input_path, args.apply, args.run_changed)
        return execute_once(input_path, args.apply, args.run_changed)
    except (SyncError, OSError) as exc:
        print(f"[sync] ERROR: {exc}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
